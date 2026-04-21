#!/usr/bin/env python3
"""
bugs_to_xlsx.py —— 把主动探索得到的 bug 清单 YAML/JSON 转成规范 Excel 汇总表

专为 qa-engineer skill 的"主动探索测试"工作流设计。输入结构见
references/exploratory-testing.md 第 6 节。

用法：
    python bugs_to_xlsx.py <bugs.yaml|bugs.json> [-o 输出.xlsx]

依赖：
    pip install openpyxl pyyaml
"""

from __future__ import annotations

import argparse
import json
import pathlib
import sys
from typing import Any


try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    print("请先安装依赖：pip install openpyxl pyyaml", file=sys.stderr)
    raise


COLUMNS = [
    ("id", "Bug ID", 12),
    ("title", "标题（一句话结论）", 45),
    ("level", "档位", 8),
    ("platform", "平台", 10),
    ("category", "分类", 10),
    ("module", "模块", 15),
    ("severity", "严重度", 10),
    ("priority", "优先级", 10),
    ("status", "状态", 12),
    ("reproducible", "必现/偶现", 12),
    ("steps", "复现步骤", 40),
    ("expected", "期望结果", 30),
    ("actual", "实际结果", 30),
    ("evidence", "证据", 30),
    ("suggestion", "修复建议", 30),
    ("tags", "标签", 15),
]

HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="C0504D")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="微软雅黑", size=10)
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)

THIN = Side(border_style="thin", color="D9D9D9")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

SEVERITY_FILL = {
    "S1": PatternFill("solid", fgColor="FFC7CE"),
    "S2": PatternFill("solid", fgColor="FFEB9C"),
    "S3": PatternFill("solid", fgColor="FFF2CC"),
    "S4": PatternFill("solid", fgColor="DDEBF7"),
}
PRIORITY_FILL = {
    "P0": PatternFill("solid", fgColor="FFC7CE"),
    "P1": PatternFill("solid", fgColor="FFEB9C"),
    "P2": PatternFill("solid", fgColor="FFF2CC"),
    "P3": PatternFill("solid", fgColor="DDEBF7"),
}
STATUS_FILL = {
    "已确认": PatternFill("solid", fgColor="C6EFCE"),
    "待验证": PatternFill("solid", fgColor="FFEB9C"),
}


def load_bugs(path: str) -> dict[str, Any]:
    text = pathlib.Path(path).read_text(encoding="utf-8")
    if path.endswith((".yaml", ".yml")):
        import yaml
        return yaml.safe_load(text)
    return json.loads(text)


def _flatten_numbered(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return "\n".join(f"{i+1}. {v}" for i, v in enumerate(value))
    return str(value)


def _flatten_plain(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return "\n".join(str(x) for x in value)
    return str(value)


def _flatten_evidence(value: Any) -> str:
    """evidence 列表每项是 dict，拼成可读字符串。"""
    if value is None:
        return ""
    if not isinstance(value, (list, tuple)):
        return str(value)
    parts: list[str] = []
    for i, item in enumerate(value, start=1):
        if isinstance(item, dict):
            t = item.get("type", "?")
            detail = item.get("path") or item.get("content") or item.get("snippet") or ""
            parts.append(f"{i}. [{t}] {detail}")
        else:
            parts.append(f"{i}. {item}")
    return "\n".join(parts)


def build_workbook(data: dict[str, Any]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    project = data.get("project", "Bug 汇总")
    ws.title = project[:30]

    # 表头
    for col_idx, (key, header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28

    bugs = data.get("bugs", [])
    for row_idx, bug in enumerate(bugs, start=2):
        for col_idx, (key, _, _) in enumerate(COLUMNS, start=1):
            raw = bug.get(key)
            if key == "steps":
                value = _flatten_numbered(raw)
            elif key in ("expected", "actual", "tags"):
                value = _flatten_plain(raw)
            elif key == "evidence":
                value = _flatten_evidence(raw)
            else:
                value = "" if raw is None else str(raw)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = BORDER
            if key == "severity" and value in SEVERITY_FILL:
                cell.fill = SEVERITY_FILL[value]
            elif key == "priority" and value in PRIORITY_FILL:
                cell.fill = PRIORITY_FILL[value]
            elif key == "status" and value in STATUS_FILL:
                cell.fill = STATUS_FILL[value]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(len(bugs)+1, 1)}"

    # 下拉验证
    if bugs:
        last_row = len(bugs) + 1
        for key, formula in [
            ("severity", '"S1,S2,S3,S4"'),
            ("priority", '"P0,P1,P2,P3"'),
            ("status", '"已确认,待验证"'),
            ("reproducible", '"必现,偶现,一次"'),
            ("level", '"L0,L1,L2,L3,L4"'),
            ("platform", '"pc,mobile,both"'),
        ]:
            col = next((i for i, (k, _, _) in enumerate(COLUMNS, start=1) if k == key), None)
            if not col:
                continue
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.add(f"{get_column_letter(col)}2:{get_column_letter(col)}{last_row}")
            ws.add_data_validation(dv)

        # S1/P0 整行不太容易，退而求其次：在列上加条件格式
        sev_col = next((i for i, (k, _, _) in enumerate(COLUMNS, start=1) if k == "severity"), None)
        if sev_col:
            sev_range = f"{get_column_letter(sev_col)}2:{get_column_letter(sev_col)}{last_row}"
            red = PatternFill("solid", fgColor="FFC7CE")
            ws.conditional_formatting.add(
                sev_range, CellIsRule(operator="equal", formula=['"S1"'], fill=red)
            )

    # 底部元信息
    if bugs:
        meta_row = len(bugs) + 3
        for i, (label, val_key) in enumerate([
            ("项目：", "project"),
            ("被测对象：", "target"),
            ("探索模式：", "mode"),
            ("扫描时间：", "scanned_at"),
            ("Bug 总数：", None),
        ]):
            ws.cell(row=meta_row + i, column=1, value=label).font = Font(bold=True)
            if val_key is None:
                ws.cell(row=meta_row + i, column=2, value=len(bugs))
            else:
                ws.cell(row=meta_row + i, column=2, value=str(data.get(val_key, "")))

    return wb


def main() -> int:
    parser = argparse.ArgumentParser(description="把 bug 清单转成规范 Excel 汇总表")
    parser.add_argument("input", help="输入：bugs.yaml 或 bugs.json")
    parser.add_argument("-o", "--output", default=None, help="输出 .xlsx；默认 <project>-bugs.xlsx")
    args = parser.parse_args()

    data = load_bugs(args.input)
    wb = build_workbook(data)

    if args.output:
        out_path = pathlib.Path(args.output)
    else:
        project = data.get("project", "bugs")
        out_path = pathlib.Path(f"{project}-bugs.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    print(f"已生成：{out_path}")
    print(f"Bug 数：{len(data.get('bugs', []))}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
