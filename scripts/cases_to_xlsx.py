#!/usr/bin/env python3
"""
cases_to_xlsx.py —— 把结构化用例数据转成规范的 Excel 文件

支持：
- 表头冻结、加粗、填色
- 自适应列宽
- 优先级 / 结果列下拉验证
- Pass=绿 / Fail=红 条件格式
- 标准格式 + 禅道导入格式

用法：
    python cases_to_xlsx.py <cases.yaml|cases.json> [-o 输出文件.xlsx] [--format standard|zentao]

输入格式（YAML 示例）：
    project: 订单模块测试
    version: v1.2.0
    cases:
      - id: TC-ORDER-001
        module: 订单创建
        sub_module: 正常下单
        priority: P0
        title: 使用默认地址下单成功
        preconditions: [已登录, 购物车有商品]
        steps: [进入购物车, 结算, 提交]
        expected: [跳转支付页, 订单号存在]
        case_type: 功能测试
        design_method: 场景法
        tags: [smoke]
        related_req: PRD-456

依赖：
    pip install openpyxl pyyaml
"""

from __future__ import annotations

import argparse
import json
import pathlib
import sys
from typing import Any


try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    print("请先安装依赖：pip install openpyxl pyyaml", file=sys.stderr)
    raise


# ---------- 列定义 ----------

STANDARD_COLUMNS = [
    ("id", "用例 ID", 18),
    ("module", "模块", 15),
    ("sub_module", "子模块", 15),
    ("priority", "优先级", 10),
    ("title", "用例标题", 40),
    ("preconditions", "前置条件", 30),
    ("steps", "测试步骤", 40),
    ("expected", "预期结果", 30),
    ("case_type", "用例类型", 12),
    ("design_method", "设计方法", 12),
    ("tags", "Tags", 15),
    ("related_req", "关联需求", 15),
    ("actual", "实际结果", 25),
    ("executor", "执行人", 10),
    ("result", "结果", 10),
]

# 禅道官方导入模板列顺序（可按自家禅道微调）
ZENTAO_COLUMNS = [
    ("id", "编号", 10),
    ("module", "所属模块", 15),
    ("title", "用例标题", 40),
    ("preconditions", "前置条件", 30),
    ("steps", "步骤", 40),
    ("expected", "预期", 30),
    ("case_type", "用例类型", 12),
    ("priority", "优先级", 10),
    ("related_req", "相关需求", 15),
]


# ---------- 样式 ----------

HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

BODY_FONT = Font(name="微软雅黑", size=10)
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)

THIN = Side(border_style="thin", color="D9D9D9")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

PRIORITY_FILL = {
    "P0": PatternFill("solid", fgColor="FFC7CE"),
    "P1": PatternFill("solid", fgColor="FFEB9C"),
    "P2": PatternFill("solid", fgColor="FFF2CC"),
    "P3": PatternFill("solid", fgColor="DDEBF7"),
}


# ---------- 工具 ----------

def load_cases(path: str) -> dict[str, Any]:
    text = pathlib.Path(path).read_text(encoding="utf-8")
    if path.endswith((".yaml", ".yml")):
        import yaml
        return yaml.safe_load(text)
    return json.loads(text)


def _flatten(value: Any) -> str:
    """把 list 或 str 变成多行字符串。list 每项前加编号。"""
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        if all(isinstance(x, str) for x in value):
            return "\n".join(f"{i+1}. {v}" for i, v in enumerate(value))
        return "\n".join(str(x) for x in value)
    return str(value)


def _flatten_simple(value: Any) -> str:
    """不加编号的多行字符串（用于 tags / preconditions）。"""
    if value is None:
        return ""
    if isinstance(value, (list, tuple)):
        return "\n".join(str(x) for x in value)
    return str(value)


# ---------- 主逻辑 ----------

def build_workbook(data: dict[str, Any], format_name: str) -> Workbook:
    columns = ZENTAO_COLUMNS if format_name == "zentao" else STANDARD_COLUMNS
    wb = Workbook()
    ws = wb.active
    project = data.get("project", "测试用例")
    ws.title = project[:30]  # Excel sheet 名上限 31 字符

    # 写表头
    for col_idx, (key, header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 28

    # 写数据
    cases = data.get("cases", [])
    for row_idx, case in enumerate(cases, start=2):
        for col_idx, (key, _, _) in enumerate(columns, start=1):
            raw = case.get(key)
            if key in ("steps", "expected"):
                value = _flatten(raw)
            elif key in ("preconditions", "tags"):
                value = _flatten_simple(raw)
            else:
                value = "" if raw is None else str(raw)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = BORDER
            # 优先级列填色
            if key == "priority" and value in PRIORITY_FILL:
                cell.fill = PRIORITY_FILL[value]

    # 冻结首行
    ws.freeze_panes = "A2"

    # 筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(len(cases)+1, 1)}"

    # 数据验证：优先级、结果
    if cases:
        last_row = len(cases) + 1
        priority_col = next((i for i, (k, _, _) in enumerate(columns, start=1) if k == "priority"), None)
        result_col = next((i for i, (k, _, _) in enumerate(columns, start=1) if k == "result"), None)

        if priority_col:
            dv = DataValidation(type="list", formula1='"P0,P1,P2,P3"', allow_blank=True)
            dv.add(f"{get_column_letter(priority_col)}2:{get_column_letter(priority_col)}{last_row}")
            ws.add_data_validation(dv)

        if result_col:
            dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked,N/A"', allow_blank=True)
            result_range = f"{get_column_letter(result_col)}2:{get_column_letter(result_col)}{last_row}"
            dv.add(result_range)
            ws.add_data_validation(dv)

            # 条件格式：Pass=绿 / Fail=红
            green_fill = PatternFill("solid", fgColor="C6EFCE")
            red_fill = PatternFill("solid", fgColor="FFC7CE")
            ws.conditional_formatting.add(
                result_range,
                CellIsRule(operator="equal", formula=['"Pass"'], fill=green_fill),
            )
            ws.conditional_formatting.add(
                result_range,
                CellIsRule(operator="equal", formula=['"Fail"'], fill=red_fill),
            )

    # 在底部追加 metadata（只在 standard 格式）
    if format_name == "standard" and cases:
        meta_row = len(cases) + 3
        ws.cell(row=meta_row, column=1, value="项目：").font = Font(bold=True)
        ws.cell(row=meta_row, column=2, value=data.get("project", ""))
        ws.cell(row=meta_row + 1, column=1, value="版本：").font = Font(bold=True)
        ws.cell(row=meta_row + 1, column=2, value=data.get("version", ""))
        ws.cell(row=meta_row + 2, column=1, value="作者：").font = Font(bold=True)
        ws.cell(row=meta_row + 2, column=2, value=data.get("author", ""))
        ws.cell(row=meta_row + 3, column=1, value="创建时间：").font = Font(bold=True)
        ws.cell(row=meta_row + 3, column=2, value=str(data.get("created", "")))

    return wb


def main() -> int:
    parser = argparse.ArgumentParser(description="把结构化用例转成规范 Excel")
    parser.add_argument("input", help="输入文件：cases.yaml 或 cases.json")
    parser.add_argument("-o", "--output", default=None, help="输出 .xlsx 路径；默认 <project>.xlsx")
    parser.add_argument("--format", choices=["standard", "zentao"], default="standard",
                        help="输出格式：standard（默认）或 zentao（禅道导入模板）")
    args = parser.parse_args()

    data = load_cases(args.input)
    wb = build_workbook(data, args.format)

    if args.output:
        out_path = pathlib.Path(args.output)
    else:
        project = data.get("project", "testcases")
        out_path = pathlib.Path(f"{project}.xlsx")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    case_count = len(data.get("cases", []))
    print(f"已生成：{out_path}")
    print(f"用例数：{case_count}")
    print(f"格式：{args.format}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
